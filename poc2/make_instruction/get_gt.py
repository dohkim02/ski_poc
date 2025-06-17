import pandas as pd
import json
import ast
from collections import defaultdict
import numpy as np


class ExcelGroupProcessor:
    def __init__(self, file_path):
        """
        엑셀 파일을 읽어서 초기화

        Args:
            file_path (str): 엑셀 파일 경로
        """
        self.file_path = file_path
        self.df = None
        self.grouped_data = None

    def load_data(self):
        """엑셀 파일을 읽어오기"""
        try:
            # 엑셀 파일 확장자에 따라 다르게 처리
            if self.file_path.endswith(".csv"):
                self.df = pd.read_csv(self.file_path, encoding="utf-8")
            else:
                self.df = pd.read_excel(self.file_path)

            print(f"데이터 로드 완료: {len(self.df)}행, {len(self.df.columns)}열")
            print("컬럼명:", list(self.df.columns))
            return True

        except Exception as e:
            print(f"파일 로드 중 오류 발생: {e}")
            return False

    def parse_usage_pattern(self, pattern_str):
        """
        사용량_패턴 문자열을 파싱하여 딕셔너리로 변환

        Args:
            pattern_str (str): 사용량 패턴 문자열

        Returns:
            dict: 월별 사용량 딕셔너리
        """
        try:
            # 문자열이 딕셔너리 형태인 경우
            if isinstance(pattern_str, str):
                # 작은따옴표를 큰따옴표로 변경하여 JSON 파싱 가능하게 만들기
                pattern_str = pattern_str.replace("'", '"')
                return json.loads(pattern_str)
            elif isinstance(pattern_str, dict):
                return pattern_str
            else:
                return {}
        except:
            try:
                # ast.literal_eval을 사용한 파싱 시도
                return ast.literal_eval(str(pattern_str))
            except:
                print(f"사용량 패턴 파싱 실패: {pattern_str}")
                return {}

    def calculate_monthly_stats(self, usage_patterns):
        """
        여러 사용량 패턴의 월별 평균과 표준편차를 계산

        Args:
            usage_patterns (list): 사용량 패턴 딕셔너리 리스트

        Returns:
            tuple: (월별 평균 사용량, 월별 표준편차)
        """
        monthly_values = defaultdict(list)

        for pattern in usage_patterns:
            if isinstance(pattern, dict):
                for month, value in pattern.items():
                    try:
                        monthly_values[month].append(float(value))
                    except (ValueError, TypeError):
                        continue

        # 각 월별 평균과 표준편차 계산
        monthly_averages = {}
        monthly_stds = {}

        for month, values in monthly_values.items():
            if values:
                monthly_averages[month] = round(np.mean(values), 2)
                # 표준편차 계산 (모집단 표준편차 사용)
                monthly_stds[month] = (
                    round(np.std(values, ddof=0), 2) if len(values) > 1 else 0.0
                )

        return monthly_averages, monthly_stds

    def group_and_calculate(self, group_column="용도"):
        """
        지정된 컬럼을 기준으로 그룹화하고 평균값과 표준편차 계산

        Args:
            group_column (str): 그룹화할 컬럼명

        Returns:
            pd.DataFrame: 그룹화된 결과 데이터프레임
        """
        if self.df is None:
            print("데이터가 로드되지 않았습니다.")
            return None

        # 필요한 컬럼들이 존재하는지 확인
        required_columns = [group_column, "보일러 열량", "연소기 열량", "사용량_패턴"]
        missing_columns = [
            col for col in required_columns if col not in self.df.columns
        ]

        if missing_columns:
            print(f"필요한 컬럼이 없습니다: {missing_columns}")
            return None

        # 그룹별로 데이터 처리
        grouped_results = []

        for group_value, group_df in self.df.groupby(group_column):
            # 보일러 열량 평균과 표준편차
            boiler_avg = group_df["보일러 열량"].mean()
            boiler_std = (
                group_df["보일러 열량"].std(ddof=0) if len(group_df) > 1 else 0.0
            )

            # 연소기 열량 평균과 표준편차
            combustor_avg = group_df["연소기 열량"].mean()
            combustor_std = (
                group_df["연소기 열량"].std(ddof=0) if len(group_df) > 1 else 0.0
            )

            # 사용량 패턴 파싱 및 평균, 표준편차 계산
            usage_patterns = []
            for pattern in group_df["사용량_패턴"]:
                parsed_pattern = self.parse_usage_pattern(pattern)
                if parsed_pattern:
                    usage_patterns.append(parsed_pattern)

            # 월별 평균 사용량과 표준편차 계산
            monthly_averages, monthly_stds = self.calculate_monthly_stats(
                usage_patterns
            )

            # 결과 저장
            grouped_results.append(
                {
                    group_column: group_value,
                    "보일러 열량 평균": round(boiler_avg, 2),
                    "보일러 열량 표준편차": round(boiler_std, 2),
                    "연소기 열량 평균": round(combustor_avg, 2),
                    "연소기 열량 표준편차": round(combustor_std, 2),
                    "사용량 패턴 평균": monthly_averages,
                    "사용량 패턴 표준편차": monthly_stds,
                }
            )

        # 데이터프레임으로 변환
        self.grouped_data = pd.DataFrame(grouped_results)

        print(f"그룹화 완료: {len(self.grouped_data)}개 그룹")
        return self.grouped_data

    def save_to_excel(self, output_path, group_name, include_monthly_columns=True):
        """
        결과를 엑셀 파일로 저장 (평균과 표준편차 포함)

        Args:
            output_path (str): 저장할 파일 경로
            include_monthly_columns (bool): 월별 사용량을 별도 컬럼으로 분리할지 여부
        """
        if self.grouped_data is None:
            print("그룹화된 데이터가 없습니다.")
            return False

        try:
            if include_monthly_columns:
                # 월별 사용량을 별도 컬럼으로 분리
                expanded_data = []

                for _, row in self.grouped_data.iterrows():
                    # 용도 컬럼명 올바르게 가져오기
                    group_column_name = self.grouped_data.columns[0]
                    group_value = row[group_column_name]

                    new_row = {
                        group_name: group_value,
                        "보일러_열량_평균": round(row["보일러 열량 평균"], 2),
                        "보일러_열량_표준편차": round(row["보일러 열량 표준편차"], 2),
                        "연소기_열량_평균": round(row["연소기 열량 평균"], 2),
                        "연소기_열량_표준편차": round(row["연소기 열량 표준편차"], 2),
                    }

                    # 월별 평균 데이터 추가
                    monthly_avg_data = row["사용량 패턴 평균"]
                    monthly_std_data = row["사용량 패턴 표준편차"]

                    if isinstance(monthly_avg_data, dict) and isinstance(
                        monthly_std_data, dict
                    ):
                        # 월 순서대로 정렬
                        month_order = [
                            "1월",
                            "2월",
                            "3월",
                            "4월",
                            "5월",
                            "6월",
                            "7월",
                            "8월",
                            "9월",
                            "10월",
                            "11월",
                            "12월",
                        ]

                        for month in month_order:
                            # 평균값
                            if month in monthly_avg_data:
                                new_row[f"사용량_{month}_평균"] = round(
                                    monthly_avg_data[month], 2
                                )
                            else:
                                new_row[f"사용량_{month}_평균"] = 0

                            # 표준편차
                            if month in monthly_std_data:
                                new_row[f"사용량_{month}_표준편차"] = round(
                                    monthly_std_data[month], 2
                                )
                            else:
                                new_row[f"사용량_{month}_표준편차"] = 0

                    expanded_data.append(new_row)

                result_df = pd.DataFrame(expanded_data)

                # 엑셀 파일로 저장하고 서식 적용
                from openpyxl import Workbook
                from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
                from openpyxl.utils.dataframe import dataframe_to_rows

                wb = Workbook()
                ws = wb.active
                ws.title = "그룹화_결과_평균_표준편차"

                # 데이터 삽입
                for r in dataframe_to_rows(result_df, index=False, header=True):
                    ws.append(r)

                # 헤더 스타일 적용
                header_fill = PatternFill(
                    start_color="4472C4", end_color="4472C4", fill_type="solid"
                )
                header_font = Font(color="FFFFFF", bold=True, size=10)
                header_alignment = Alignment(horizontal="center", vertical="center")

                # 테두리 스타일
                thin_border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin"),
                )

                # 헤더 행 스타일 적용
                for col in range(1, len(result_df.columns) + 1):
                    cell = ws.cell(row=1, column=col)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment
                    cell.border = thin_border

                # 데이터 행 스타일 적용
                for row in range(2, len(result_df) + 2):
                    for col in range(1, len(result_df.columns) + 1):
                        cell = ws.cell(row=row, column=col)
                        cell.border = thin_border

                        # 숫자 컬럼은 우측 정렬
                        if col > 1:  # 그룹 컬럼 제외
                            cell.alignment = Alignment(
                                horizontal="right", vertical="center"
                            )
                            # 천 단위 구분 기호 적용
                            if (
                                isinstance(cell.value, (int, float))
                                and cell.value > 1000
                            ):
                                cell.number_format = "#,##0.00"
                        else:
                            cell.alignment = Alignment(
                                horizontal="center", vertical="center"
                            )

                # 열 너비 자동 조정
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter

                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass

                    adjusted_width = min(
                        max_length + 2, 25
                    )  # 최대 25로 제한 (컬럼명이 길어짐)
                    ws.column_dimensions[column_letter].width = adjusted_width

                wb.save(output_path)

            else:
                # 간단한 저장 (월별 컬럼 분리 없음)
                result_df = self.grouped_data.copy()
                result_df.to_excel(output_path, index=False, engine="openpyxl")

            print(f"결과가 저장되었습니다: {output_path}")
            print(f"총 {len(result_df)}개 그룹의 데이터가 저장되었습니다.")
            return True

        except Exception as e:
            print(f"파일 저장 중 오류 발생: {e}")
            return False

    def display_results(self):
        """결과를 화면에 출력 (평균과 표준편차 포함)"""
        if self.grouped_data is None:
            print("그룹화된 데이터가 없습니다.")
            return

        print("\n" + "=" * 100)
        print("                        그룹화 결과 (평균 ± 표준편차)")
        print("=" * 100)

        for idx, (_, row) in enumerate(self.grouped_data.iterrows(), 1):
            group_column_name = self.grouped_data.columns[0]
            group_value = row[group_column_name]

            print(f"\n[{idx}] 그룹: {group_value}")
            print("-" * 70)
            print(
                f"📊 보일러 열량: {row['보일러 열량 평균']:>12,.2f} ± {row['보일러 열량 표준편차']:>8,.2f}"
            )
            print(
                f"🔥 연소기 열량: {row['연소기 열량 평균']:>12,.2f} ± {row['연소기 열량 표준편차']:>8,.2f}"
            )
            print(f"📈 월별 사용량 (평균 ± 표준편차):")

            monthly_avg_data = row["사용량 패턴 평균"]
            monthly_std_data = row["사용량 패턴 표준편차"]

            if isinstance(monthly_avg_data, dict) and isinstance(
                monthly_std_data, dict
            ):
                # 월 순서대로 정렬하여 출력
                month_order = [
                    "1월",
                    "2월",
                    "3월",
                    "4월",
                    "5월",
                    "6월",
                    "7월",
                    "8월",
                    "9월",
                    "10월",
                    "11월",
                    "12월",
                ]

                # 분기별로 나누어 출력
                quarters = [
                    ["1월", "2월", "3월"],  # 1분기
                    ["4월", "5월", "6월"],  # 2분기
                    ["7월", "8월", "9월"],  # 3분기
                    ["10월", "11월", "12월"],  # 4분기
                ]

                for quarter_idx, quarter in enumerate(quarters, 1):
                    quarter_data = []
                    for month in quarter:
                        avg_val = monthly_avg_data.get(month, 0)
                        std_val = monthly_std_data.get(month, 0)
                        quarter_data.append(f"{month}: {avg_val:>7.2f}±{std_val:>6.2f}")

                    print(f"   {quarter_idx}분기: {' | '.join(quarter_data)}")

        print("\n" + "=" * 100)

    def get_results_as_dict(self):
        """
        그룹화된 결과를 딕셔너리 형태로 반환 (평균 ± 표준편차 포함)

        Returns:
            list: 각 그룹별 결과를 담은 딕셔너리 리스트
        """
        if self.grouped_data is None:
            print("그룹화된 데이터가 없습니다.")
            return []

        results = []

        for _, row in self.grouped_data.iterrows():
            group_column_name = self.grouped_data.columns[0]
            group_value = row[group_column_name]

            result_entry = {
                "그룹": group_value,
                "보일러 열량 평균": round(row["보일러 열량 평균"], 2),
                "보일러 열량 표준편차": round(row["보일러 열량 표준편차"], 2),
                "연소기 열량 평균": round(row["연소기 열량 평균"], 2),
                "연소기 열량 표준편차": round(row["연소기 열량 표준편차"], 2),
                "월별 사용량 평균": {},
                "월별 사용량 표준편차": {},
                "분기별 요약": {},
            }

            monthly_avg_data = row["사용량 패턴 평균"]
            monthly_std_data = row["사용량 패턴 표준편차"]

            if isinstance(monthly_avg_data, dict) and isinstance(
                monthly_std_data, dict
            ):
                month_order = [f"{i}월" for i in range(1, 13)]
                quarters = {
                    "1분기": month_order[0:3],
                    "2분기": month_order[3:6],
                    "3분기": month_order[6:9],
                    "4분기": month_order[9:12],
                }

                for month in month_order:
                    result_entry["월별 사용량 평균"][month] = round(
                        monthly_avg_data.get(month, 0), 2
                    )
                    result_entry["월별 사용량 표준편차"][month] = round(
                        monthly_std_data.get(month, 0), 2
                    )

                # 분기별 요약 추가
                for q_name, months in quarters.items():
                    quarter_data = [
                        f"{month}: {monthly_avg_data.get(month, 0):.2f}±{monthly_std_data.get(month, 0):.2f}"
                        for month in months
                    ]
                    result_entry["분기별 요약"][q_name] = " | ".join(quarter_data)

            results.append(result_entry)

        return results

    def get_outlier_detection_bounds(self, group_value, month, threshold=2):
        """
        특정 그룹과 월에 대한 이상치 탐지 범위를 반환

        Args:
            group_value: 그룹값 (예: "상업용")
            month: 월 (예: "3월")
            threshold: 표준편차 배수 (기본값: 2)

        Returns:
            dict: 이상치 탐지 정보
        """
        if self.grouped_data is None:
            return None

        # 해당 그룹 찾기
        group_column_name = self.grouped_data.columns[0]
        group_row = self.grouped_data[
            self.grouped_data[group_column_name] == group_value
        ]

        if group_row.empty:
            print(f"그룹 '{group_value}'를 찾을 수 없습니다.")
            return None

        row = group_row.iloc[0]
        monthly_avg_data = row["사용량 패턴 평균"]
        monthly_std_data = row["사용량 패턴 표준편차"]

        if month not in monthly_avg_data or month not in monthly_std_data:
            print(f"월 '{month}' 데이터를 찾을 수 없습니다.")
            return None

        mean_val = monthly_avg_data[month]
        std_val = monthly_std_data[month]

        lower_bound = mean_val - threshold * std_val
        upper_bound = mean_val + threshold * std_val

        return {
            "group": group_value,
            "month": month,
            "mean": mean_val,
            "std": std_val,
            "lower_bound": max(0, lower_bound),  # 음수 방지
            "upper_bound": upper_bound,
            "threshold": threshold,
        }


def detect_outlier(processor, new_usage, group_value, month, threshold=2):
    """
    새로운 사용량 데이터가 이상치인지 판단

    Args:
        processor: ExcelGroupProcessor 인스턴스
        new_usage: 새로운 사용량 값
        group_value: 그룹값
        month: 월
        threshold: 표준편차 배수

    Returns:
        dict: 이상치 판단 결과
    """
    bounds = processor.get_outlier_detection_bounds(group_value, month, threshold)

    if bounds is None:
        return None

    is_outlier = (new_usage < bounds["lower_bound"]) or (
        new_usage > bounds["upper_bound"]
    )
    z_score = (
        abs(new_usage - bounds["mean"]) / bounds["std"] if bounds["std"] > 0 else 0
    )

    return {
        "new_usage": new_usage,
        "is_outlier": is_outlier,
        "z_score": z_score,
        "bounds": bounds,
        "message": f"{'이상치' if is_outlier else '정상'} (Z-score: {z_score:.2f})",
    }


def convert_df_to_dict(obj):
    if isinstance(obj, pd.DataFrame):
        return obj.to_dict()
    elif isinstance(obj, list):
        return [convert_df_to_dict(item) for item in obj]
    elif isinstance(obj, dict):
        return {k: convert_df_to_dict(v) for k, v in obj.items()}
    else:
        return obj


def main(input_file, output_file, group_name, gt_json_path):
    """메인 실행 함수"""
    # 프로세서 초기화
    processor = ExcelGroupProcessor(input_file)

    # 데이터 로드
    if not processor.load_data():
        return

    # 그룹화 및 평균/표준편차 계산
    result = processor.group_and_calculate(group_column=group_name)
    results_serializable = convert_df_to_dict(result)
    # JSON 파일로 저장

    with open(gt_json_path, "w", encoding="utf-8") as f:
        json.dump(results_serializable, f, ensure_ascii=False, indent=4)

    print(f"JSON 파일로 저장 완료: {gt_json_path}")
    # if result is not None:
    #     # 결과 출력
    #     # processor.display_results()
    #     res_dict = processor.get_results_as_dict()
    #     print(res_dict)
    #     # 엑셀 파일로 저장 (월별 컬럼 분리, 평균과 표준편차 포함)
    #     processor.save_to_excel(output_file, group_name, include_monthly_columns=True)

    #     # 이상치 탐지 예시
    #     print("\n" + "=" * 50)
    #     print("이상치 탐지 예시")
    #     print("=" * 50)

    #     # 예시: 첫 번째 그룹의 1월 데이터로 이상치 탐지 테스트
    #     if len(processor.grouped_data) > 0:
    #         first_group = processor.grouped_data.iloc[0][group_name]
    #         test_usage = 1500  # 테스트용 사용량
    #         result = detect_outlier(
    #             processor, test_usage, first_group, "1월", threshold=2
    #         )

    #         if result:
    #             print(f"테스트 결과: {result['message']}")
    #             print(f"입력값: {result['new_usage']}")
    #             print(
    #                 f"정상 범위: {result['bounds']['lower_bound']:.2f} ~ {result['bounds']['upper_bound']:.2f}"
    #             )

    return processor


if __name__ == "__main__":
    # 실제 데이터로 실행하려면 main() 함수 사용
    input_file = "./data2_preprocessed2.xlsx"
    output_file = "./data2_biz_with_std.xlsx"
    gt_json_path = "./biz_group_gt.json"
    group_name = "그룹"
    processor = main(input_file, output_file, group_name, gt_json_path)
